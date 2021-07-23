from openpyxl import load_workbook
from kivy.app import App
from kivy.uix.widget import Widget
from kivy.properties import ObjectProperty
from kivy.lang import Builder
from kivy.config import Config
Config.set('graphics', 'resizable', '0')
Config.set('graphics', 'width', '500')
Config.set('graphics', 'height', '200')

Builder.load_file("layout.kv")


class MyGridLayout(Widget):
    read_file_name = ObjectProperty(None)
    column = ObjectProperty(None)
    write_file_name = ObjectProperty(None)

    def press(self):

        read_file_name = self.read_file_name.text + '.xlsx'
        column = self.column.text
        write_file_name = self.write_file_name.text+ '.xlsx'
        print(f'Your read file name  {read_file_name}  Your column name:  {column}  and output file name:  {write_file_name}')

        # Clear the input boxes
        self.read_file_name.text = ""
        self.column.text = ""
        self.write_file_name.text = ""

        # ----------------------- Logic below
        file_name = read_file_name

        # Read excel sheet
        current_book = load_workbook(file_name)
        sheet = current_book.active

        # Variables
        total_rows = sheet.max_row
        decimal_values = []
        binary_values = []
        column = column
        output_file_name = write_file_name

        # -------------------------------------------------
        # Read data from column and save in list
        for i in range(2, total_rows + 1, 1):
            a = i
            cell = column + a.__str__()
            if sheet[cell].value is None:
                continue
            decimal_values.append(sheet[cell].value)

        # -------------------------------------------------
        # Convert decimal value in binary
        for i in range(decimal_values.__len__()):
            binary_values.append(bin(decimal_values[i]))

        # -------------------------------------------------
        # # Split binary values in single digit and create list
        # split_binary = []
        #
        # print(binary_values.__len__())
        # print(split_binary)
        # for i in range(binary_values.__len__()):
        #     n = 1
        #     split_binary.append([binary_values[i][j:j + n] for j in range(0, len(binary_values[i]), n)])
        #
        # print(split_binary)

        # ------------------------------------------------------------------------------------
        #  Find location of 1 in binary number and append in new list by separating with comma
        final_output = []
        for i in range(binary_values.__len__()):
            port = ''
            for j in range(binary_values[i].__len__() - 1, 0, -1):
                if binary_values[i][j] == '1':
                    port += (binary_values[i].__len__() - j).__str__()
                    port += ','
            final_output.append(port)

        # --------------------------------------------------------------------------------------
        # do changes in sheet
        for i in range(final_output.__len__()):
            a = i + 2
            cell = column + a.__str__()
            sheet[cell] = final_output[i]

        # save changes in new file
        current_book.save(output_file_name)


class CrosspointDecoderApp(App):
    def build(self):
        return MyGridLayout()


if __name__ == '__main__':
    CrosspointDecoderApp().run()

